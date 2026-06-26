"""
Generate tools/template.xlsx (and template.csv) — the blank form admins fill in
to add a new exam. Re-run after changing the columns here.

Usage:
    python tools/make_template.py
"""
import csv
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))

HEADERS = ["exam_id", "exam_name", "group", "question",
           "A", "B", "C", "D", "answer", "explanation", "tags"]

EXAMPLE = [
    ["reading-1", "Reading 1 — Demo", "Chương 1",
     "2 + 2 bằng mấy?", "3", "4", "5", "", "B",
     "2 cộng 2 bằng 4.", "demo;math"],
    ["reading-1", "Reading 1 — Demo", "Chương 1",
     "Thủ đô Việt Nam là?", "TP.HCM", "Đà Nẵng", "Hà Nội", "", "C",
     "Thủ đô là Hà Nội.", "demo"],
]

NOTES = [
    "HƯỚNG DẪN",
    "1. Mỗi dòng = 1 câu hỏi. Các câu cùng 'exam_id' gộp thành 1 đề.",
    "2. Cột bắt buộc: exam_id, question, A, B, answer.",
    "3. C, D để trống nếu câu chỉ có 2-3 đáp án.",
    "4. 'answer' điền đúng 1 chữ: A/B/C/D.",
    "5. 'tags' ngăn cách bằng ; hoặc , (tuỳ chọn).",
    "6. Lưu file rồi chạy:",
    "      python tools/xlsx_to_json.py <file.xlsx> --subject <ten_mon>",
    "   ví dụ: python tools/xlsx_to_json.py de_kinh_te.xlsx --subject economics",
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"

    header_fill = PatternFill("solid", fgColor="1F6FEB")
    header_font = Font(bold=True, color="FFFFFF")
    for c, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = header_fill
        cell.font = header_font
    for r, row in enumerate(EXAMPLE, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    widths = [12, 18, 12, 46, 16, 16, 16, 10, 8, 40, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    ws.freeze_panes = "A2"

    guide = wb.create_sheet("HƯỚNG DẪN")
    for r, line in enumerate(NOTES, 1):
        cell = guide.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True, size=13)
    guide.column_dimensions["A"].width = 90

    xlsx_path = os.path.join(HERE, "template.xlsx")
    wb.save(xlsx_path)
    print(f"✅ {xlsx_path}")

    csv_path = os.path.join(HERE, "template.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(EXAMPLE)
    print(f"✅ {csv_path}")


if __name__ == "__main__":
    main()
